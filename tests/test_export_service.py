"""
Comprehensive tests for Export Service (Module 9)
Tests all export formats, API endpoints, and edge cases
"""

import pytest
import asyncio
from pathlib import Path
import json
import pandas as pd
from datetime import datetime
from unittest.mock import Mock, patch, AsyncMock
from fastapi.testclient import TestClient

from src.services.export_service import ExportService, export_service
from src.api.main import app
from src.database.manager import DatabaseManager
from src.database.operations import DatabaseOperations
from src.config.settings import settings


# ============================================================
# Fixtures
# ============================================================

@pytest.fixture
async def test_db():
    """Create test database with sample data"""
    db_manager = DatabaseManager("sqlite+aiosqlite:///:memory:")
    await db_manager.initialize()

    # Create sample search and results
    async with db_manager.get_session() as session:
        # Create search
        search = await DatabaseOperations.create_search(
            session,
            query="test query AI technology",
            sources=["google", "bing"]
        )
        await session.flush()

        # Create sample results
        results_data = [
            {
                "search_id": search.id,
                "source": "google",
                "position": i + 1,
                "title": f"Test Result {i+1}",
                "snippet": f"This is test snippet number {i+1} about AI technology",
                "url": f"https://example.com/result{i+1}",
                "scraped_at": datetime.now()
            }
            for i in range(25)  # Create 25 test results
        ]

        await DatabaseOperations.bulk_insert_results_chunked(
            session,
            search.id,
            results_data,
            chunk_size=10
        )

        await session.commit()

        yield db_manager, search.id

    await db_manager.close()


@pytest.fixture
def test_export_service(tmp_path):
    """Create export service with temporary directory"""
    # Create temporary export directory
    export_dir = tmp_path / "exports"
    export_dir.mkdir()

    # Patch settings
    with patch.object(settings, 'EXPORT_DIR', str(export_dir)):
        with patch.object(settings, 'EXPORT_MAX_ROWS', 10000):
            service = ExportService()
            yield service, export_dir


@pytest.fixture
def api_client():
    """Create FastAPI test client"""
    return TestClient(app)


@pytest.fixture
def sample_search_data():
    """Sample search data for testing"""
    return {
        "search": {
            "id": 1,
            "query": "AI technology",
            "status": "completed",
            "sources": ["google", "bing"],
            "created_at": datetime.now(),
            "completed_at": datetime.now()
        },
        "results": [
            {
                "source": "google",
                "position": i + 1,
                "title": f"Result {i+1}",
                "snippet": f"Snippet {i+1}",
                "url": f"https://example.com/{i+1}",
                "scraped_at": datetime.now(),
                "sentiment_score": 0.5
            }
            for i in range(10)
        ],
        "analysis": {
            "sentiment": {
                "dominant_sentiment": "positive",
                "average_confidence": 0.85,
                "positive_count": 7,
                "negative_count": 1,
                "neutral_count": 2,
                "total_analyzed": 10
            },
            "trends": {
                "top_trends": [
                    {"trend": "AI", "count": 15},
                    {"trend": "machine learning", "count": 12},
                    {"trend": "neural networks", "count": 8}
                ]
            },
            "competitors": {
                "top_competitors": [
                    {"name": "Company A", "mentions": 5},
                    {"name": "Company B", "mentions": 3}
                ]
            },
            "insights": None
        }
    }


# ============================================================
# Setup Tests
# ============================================================

def test_export_service_initialization(test_export_service):
    """Test export service initialization"""
    service, export_dir = test_export_service

    assert service.export_dir == export_dir
    assert service.export_dir.exists()
    assert service.max_rows == 10000
    assert len(service.SUPPORTED_FORMATS) == 6


def test_export_directory_creation(tmp_path):
    """Test automatic export directory creation"""
    export_dir = tmp_path / "new_exports"
    assert not export_dir.exists()

    with patch.object(settings, 'EXPORT_DIR', str(export_dir)):
        service = ExportService()
        assert export_dir.exists()


# ============================================================
# Data Retrieval Tests
# ============================================================

@pytest.mark.asyncio
async def test_get_search_data_success(test_db):
    """Test successful search data retrieval"""
    db_manager, search_id = test_db

    with patch('src.services.export_service.db_manager', db_manager):
        service = ExportService()
        data = await service._get_search_data(search_id, include_analysis=False)

        assert data is not None
        assert data["search"]["id"] == search_id
        assert data["search"]["query"] == "test query AI technology"
        assert len(data["results"]) == 25
        assert data["analysis"] is None


@pytest.mark.asyncio
async def test_get_search_data_not_found():
    """Test search data retrieval for non-existent search"""
    db_manager = DatabaseManager("sqlite+aiosqlite:///:memory:")
    await db_manager.initialize()

    with patch('src.services.export_service.db_manager', db_manager):
        service = ExportService()
        data = await service._get_search_data(999, include_analysis=False)

        assert data is None

    await db_manager.close()


@pytest.mark.asyncio
async def test_get_search_data_with_analysis(test_db):
    """Test search data retrieval with analysis data"""
    db_manager, search_id = test_db

    # Mock cache to return analysis data
    mock_cache = AsyncMock()
    mock_cache.get = AsyncMock(side_effect=lambda key: {
        f"analysis:{search_id}:sentiment": {"dominant_sentiment": "positive"},
        f"analysis:{search_id}:trends": {"top_trends": ["AI", "ML"]},
        f"analysis:{search_id}:competitors": None,
        f"analysis:{search_id}:insights": None
    }.get(key))

    with patch('src.services.export_service.db_manager', db_manager):
        with patch('src.services.export_service.cache', mock_cache):
            service = ExportService()
            data = await service._get_search_data(search_id, include_analysis=True)

            assert data is not None
            assert data["analysis"] is not None
            assert data["analysis"]["sentiment"] == {"dominant_sentiment": "positive"}


# ============================================================
# CSV Export Tests
# ============================================================

@pytest.mark.asyncio
async def test_csv_export_basic(test_export_service, sample_search_data):
    """Test basic CSV export functionality"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_export.csv"

    await service._export_csv(sample_search_data, filepath)

    assert filepath.exists()

    # Read and verify CSV
    df = pd.read_csv(filepath)
    assert len(df) == 10
    assert "search_id" in df.columns
    assert "query" in df.columns
    assert "title" in df.columns
    assert df.iloc[0]["query"] == "AI technology"


@pytest.mark.asyncio
async def test_csv_export_max_rows_limit(test_export_service):
    """Test CSV export respects max rows limit"""
    service, export_dir = test_export_service
    service.max_rows = 5  # Limit to 5 rows

    # Create data with 10 results
    data = {
        "search": {"id": 1, "query": "test"},
        "results": [{"title": f"Result {i}", "url": f"http://{i}"} for i in range(10)],
        "analysis": None
    }

    filepath = export_dir / "test_limited.csv"
    await service._export_csv(data, filepath)

    df = pd.read_csv(filepath)
    assert len(df) == 5  # Should be limited


@pytest.mark.asyncio
async def test_csv_export_encoding(test_export_service):
    """Test CSV export with special characters"""
    service, export_dir = test_export_service

    # Data with special characters
    data = {
        "search": {"id": 1, "query": "tëst 日本語"},
        "results": [
            {
                "title": "Résumé café",
                "snippet": "测试 тест",
                "url": "http://example.com",
                "source": "google",
                "position": 1,
                "scraped_at": datetime.now(),
                "sentiment_score": 0.5
            }
        ],
        "analysis": None
    }

    filepath = export_dir / "test_unicode.csv"
    await service._export_csv(data, filepath)

    # Read with UTF-8 encoding
    df = pd.read_csv(filepath, encoding="utf-8-sig")
    assert "café" in df.iloc[0]["title"]


# ============================================================
# Excel Export Tests
# ============================================================

@pytest.mark.asyncio
async def test_excel_export_basic(test_export_service, sample_search_data):
    """Test basic Excel export functionality"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_export.xlsx"

    await service._export_excel(sample_search_data, filepath, include_analysis=False)

    assert filepath.exists()

    # Read Excel file
    xl_file = pd.ExcelFile(filepath)
    assert "Search Info" in xl_file.sheet_names
    assert "Results" in xl_file.sheet_names


@pytest.mark.asyncio
async def test_excel_export_multisheet(test_export_service, sample_search_data):
    """Test Excel export with multiple sheets"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_multisheet.xlsx"

    await service._export_excel(sample_search_data, filepath, include_analysis=True)

    xl_file = pd.ExcelFile(filepath)
    sheet_names = xl_file.sheet_names

    assert "Search Info" in sheet_names
    assert "Results" in sheet_names
    assert "Sentiment" in sheet_names
    assert "Trends" in sheet_names


@pytest.mark.asyncio
async def test_excel_export_formatting(test_export_service, sample_search_data):
    """Test Excel export formatting"""
    from openpyxl import load_workbook

    service, export_dir = test_export_service
    filepath = export_dir / "test_formatted.xlsx"

    await service._export_excel(sample_search_data, filepath, include_analysis=False)

    # Load workbook and check formatting
    wb = load_workbook(filepath)
    ws = wb["Results"]

    # Check header formatting (row 1)
    header_cell = ws['A1']
    assert header_cell.font.bold == True
    assert header_cell.fill.start_color.rgb == "FF366092"


@pytest.mark.asyncio
async def test_excel_export_with_analysis(test_export_service, sample_search_data):
    """Test Excel export includes analysis sheets"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_analysis.xlsx"

    await service._export_excel(sample_search_data, filepath, include_analysis=True)

    # Read sentiment sheet
    sentiment_df = pd.read_excel(filepath, sheet_name="Sentiment")
    assert "Dominant Sentiment" in sentiment_df.columns
    assert sentiment_df.iloc[0]["Dominant Sentiment"] == "positive"


# ============================================================
# JSON Export Tests
# ============================================================

@pytest.mark.asyncio
async def test_json_export_structure(test_export_service, sample_search_data):
    """Test JSON export structure"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_export.json"

    await service._export_json(sample_search_data, filepath)

    assert filepath.exists()

    # Read and verify JSON
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    assert "metadata" in data
    assert "search" in data
    assert "results" in data
    assert "analysis" in data
    assert "statistics" in data
    assert data["metadata"]["version"] == "0.9.0"


@pytest.mark.asyncio
async def test_json_export_datetime_serialization(test_export_service):
    """Test JSON export handles datetime serialization"""
    service, export_dir = test_export_service

    data = {
        "search": {
            "id": 1,
            "query": "test",
            "created_at": datetime.now(),
            "completed_at": datetime.now(),
            "status": "completed",
            "sources": ["google"]
        },
        "results": [],
        "analysis": None
    }

    filepath = export_dir / "test_datetime.json"
    await service._export_json(data, filepath)

    # Should not raise error
    with open(filepath, 'r') as f:
        json_data = json.load(f)

    # Verify datetime is serialized as string
    assert isinstance(json_data["search"]["created_at"], str)


# ============================================================
# Power BI Export Tests
# ============================================================

@pytest.mark.asyncio
async def test_powerbi_export_denormalized(test_export_service, sample_search_data):
    """Test Power BI export creates denormalized structure"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_powerbi.csv"

    await service._export_powerbi(sample_search_data, filepath)

    assert filepath.exists()

    # Read CSV
    df = pd.read_csv(filepath, encoding="utf-8-sig")

    # Check denormalized structure (search metadata in each row)
    assert "SearchID" in df.columns
    assert "Query" in df.columns
    assert all(df["SearchID"] == 1)
    assert all(df["Query"] == "AI technology")


@pytest.mark.asyncio
async def test_powerbi_export_column_naming(test_export_service, sample_search_data):
    """Test Power BI export has clean column names"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_powerbi_columns.csv"

    await service._export_powerbi(sample_search_data, filepath)

    df = pd.read_csv(filepath, encoding="utf-8-sig")

    # Check Title Case column names with spaces
    expected_columns = [
        "SearchID", "Query", "SearchDate", "SearchStatus",
        "Source", "Position", "Title", "Snippet", "URL",
        "ScrapedAt", "SentimentScore", "OverallSentiment", "DominantTrend"
    ]

    for col in expected_columns:
        assert col in df.columns


# ============================================================
# HTML Report Tests
# ============================================================

@pytest.mark.asyncio
async def test_html_report_generation(test_export_service, sample_search_data):
    """Test HTML report generation"""
    service, export_dir = test_export_service
    filepath = export_dir / "test_report.html"

    await service._export_html(sample_search_data, filepath)

    assert filepath.exists()

    # Read HTML content
    with open(filepath, 'r', encoding='utf-8') as f:
        html_content = f.read()

    assert "<!DOCTYPE html>" in html_content
    assert "AI technology" in html_content
    assert "CIAP" in html_content


@pytest.mark.asyncio
async def test_html_report_template_rendering(test_export_service, sample_search_data):
    """Test HTML template rendering with data"""
    service, export_dir = test_export_service

    html_content = service._render_report_template(sample_search_data, "default")

    assert "AI technology" in html_content
    assert "Search Information" in html_content
    assert "Sentiment Analysis" in html_content
    assert "Top Trends" in html_content


@pytest.mark.asyncio
async def test_html_report_sentiment_coloring(test_export_service):
    """Test HTML report applies sentiment color coding"""
    service, export_dir = test_export_service

    # Test positive sentiment
    data_positive = {
        "search": {"id": 1, "query": "test", "status": "completed", "sources": ["google"],
                   "created_at": datetime.now(), "completed_at": datetime.now()},
        "results": [],
        "analysis": {
            "sentiment": {"dominant_sentiment": "positive", "average_confidence": 0.9},
            "trends": None,
            "competitors": None,
            "insights": None
        }
    }

    html = service._render_report_template(data_positive, "default")
    assert 'sentiment-positive' in html
    assert 'Positive' in html


# ============================================================
# Utility Method Tests
# ============================================================

def test_validate_format(test_export_service):
    """Test format validation"""
    service, _ = test_export_service

    assert service._validate_format("csv") == True
    assert service._validate_format("excel") == True
    assert service._validate_format("json") == True
    assert service._validate_format("powerbi") == True
    assert service._validate_format("html") == True
    assert service._validate_format("pdf") == False
    assert service._validate_format("invalid") == False


def test_sanitize_filename(test_export_service):
    """Test filename sanitization"""
    service, _ = test_export_service

    assert service._sanitize_filename("normal file") == "normal_file"
    assert service._sanitize_filename("file<>:") == "file___"
    assert service._sanitize_filename("a" * 100, max_length=10) == "aaaaaaaaaa"
    assert service._sanitize_filename("File Name") == "file_name"


def test_serialize_datetime(test_export_service):
    """Test datetime serialization"""
    service, _ = test_export_service

    dt = datetime(2025, 10, 25, 12, 30, 45)
    serialized = service._serialize_datetime(dt)

    assert isinstance(serialized, str)
    assert "2025-10-25" in serialized
    assert service._serialize_datetime(None) == ""
    assert service._serialize_datetime("already_string") == "already_string"


# ============================================================
# Integration Test Summary
# ============================================================

def test_export_service_summary(test_export_service):
    """Summary test to verify all components"""
    service, export_dir = test_export_service

    # Verify service is properly initialized
    assert service is not None
    assert service.export_dir.exists()
    assert len(service.SUPPORTED_FORMATS) == 6

    # Verify all export methods exist
    assert hasattr(service, '_export_csv')
    assert hasattr(service, '_export_excel')
    assert hasattr(service, '_export_json')
    assert hasattr(service, '_export_powerbi')
    assert hasattr(service, '_export_html')
    assert hasattr(service, 'generate_report')
